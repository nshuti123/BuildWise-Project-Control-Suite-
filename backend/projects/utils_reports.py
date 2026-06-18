import io
import html
from datetime import datetime
from reportlab.lib import colors
from reportlab.lib.pagesizes import letter
from reportlab.platypus import SimpleDocTemplate, Paragraph, Spacer, Table, TableStyle
from reportlab.lib.styles import getSampleStyleSheet, ParagraphStyle
from reportlab.lib.enums import TA_CENTER, TA_RIGHT, TA_LEFT

from .report_branding import (
    append_report_footer,
    build_header_project_info_html,
    header_project_info_style,
)


def _meta_kwargs(exported_by=None, verification_token=None, verify_url=None):
    return {
        'exported_by': exported_by,
        'verification_token': verification_token,
        'verify_url': verify_url,
    }


def generate_financial_report_pdf(project, budget_summary, exported_by=None, verification_token=None, verify_url=None):
    meta = _meta_kwargs(exported_by, verification_token, verify_url)
    buffer = io.BytesIO()
    
    # Create the PDF document with tighter margins to fit more data beautifully
    doc = SimpleDocTemplate(
        buffer, 
        pagesize=letter, 
        rightMargin=40, 
        leftMargin=40, 
        topMargin=40, 
        bottomMargin=40
    )
    elements = []
    
    # Define Custom Styles
    styles = getSampleStyleSheet()
    
    header_title_style = ParagraphStyle(
        'HeaderTitle',
        parent=styles['Heading1'],
        fontName='Helvetica-Bold',
        fontSize=24,
        textColor=colors.whitesmoke,
        alignment=TA_LEFT,
        spaceAfter=0
    )
    
    header_subtitle_style = ParagraphStyle(
        'HeaderSubtitle',
        parent=styles['Normal'],
        fontName='Helvetica',
        fontSize=12,
        textColor=colors.HexColor("#cbd5e1"),
        alignment=TA_LEFT,
        spaceAfter=0
    )
    
    section_title_style = ParagraphStyle(
        'SectionTitle',
        parent=styles['Heading2'],
        fontName='Helvetica-Bold',
        fontSize=16,
        textColor=colors.HexColor("#1e293b"),
        spaceAfter=10,
        spaceBefore=20
    )
    
    normal_style = styles['Normal']
    
    title_para = Paragraph("<b>BuildWise</b>", header_title_style)
    subtitle_para = Paragraph("Financial & Cost Control Report", header_subtitle_style)
    project_info = Paragraph(
        build_header_project_info_html(project, exported_by=exported_by),
        header_project_info_style(),
    )
    
    header_data = [[
        [title_para, Spacer(1, 4), subtitle_para], 
        project_info
    ]]
    
    # Header Table: 100% width (approx 530 points on letter paper with 40pt margins)
    header_table = Table(header_data, colWidths=[330, 200])
    header_table.setStyle(TableStyle([
        ('BACKGROUND', (0, 0), (-1, -1), colors.HexColor("#2563eb")), # Premium Blue
        ('VALIGN', (0, 0), (-1, -1), 'MIDDLE'),
        ('TOPPADDING', (0, 0), (-1, -1), 20),
        ('BOTTOMPADDING', (0, 0), (-1, -1), 20),
        ('LEFTPADDING', (0, 0), (0, 0), 20),
        ('RIGHTPADDING', (1, 0), (1, 0), 20),
    ]))
    
    elements.append(header_table)
    elements.append(Spacer(1, 30))
    
    # ---------------------------------------------------------
    # 2. EXECUTIVE SUMMARY
    # ---------------------------------------------------------
    elements.append(Paragraph("Executive Summary", section_title_style))
    
    total_planned = budget_summary.get('total_planned', 0)
    total_actual = budget_summary.get('total_actual', 0)
    variance = budget_summary.get('variance', 0)
    remaining = total_planned - total_actual
    
    # We will lay out the summary as 4 nice "cards" in a single row
    summary_data = [
        ["TOTAL BUDGET", "ACTUAL SPEND", "REMAINING", "VARIANCE"],
        [
            f"Rwf {total_planned:,.2f}", 
            f"Rwf {total_actual:,.2f}", 
            f"Rwf {remaining:,.2f}", 
            f"Rwf {variance:,.2f}"
        ]
    ]
    
    # 530 width / 4 = ~132 per col
    summary_table = Table(summary_data, colWidths=[132, 132, 132, 132])
    summary_table.setStyle(TableStyle([
        ('BACKGROUND', (0, 0), (-1, 0), colors.HexColor("#f1f5f9")),
        ('TEXTCOLOR', (0, 0), (-1, 0), colors.HexColor("#64748b")),
        ('FONTNAME', (0, 0), (-1, 0), 'Helvetica-Bold'),
        ('FONTSIZE', (0, 0), (-1, 0), 9),
        ('ALIGN', (0, 0), (-1, -1), 'CENTER'),
        ('VALIGN', (0, 0), (-1, -1), 'MIDDLE'),
        ('BOTTOMPADDING', (0, 0), (-1, 0), 8),
        ('TOPPADDING', (0, 0), (-1, 0), 8),
        
        ('FONTNAME', (0, 1), (-1, 1), 'Helvetica-Bold'),
        ('FONTSIZE', (0, 1), (-1, 1), 12),
        ('TEXTCOLOR', (0, 1), (-1, 1), colors.HexColor("#0f172a")),
        ('BOTTOMPADDING', (0, 1), (-1, 1), 15),
        ('TOPPADDING', (0, 1), (-1, 1), 10),
        
        # Draw a line below the cards
        ('LINEBELOW', (0, 1), (-1, 1), 2, colors.HexColor("#e2e8f0")),
    ]))
    
    elements.append(summary_table)
    elements.append(Spacer(1, 30))
    
    # ---------------------------------------------------------
    # 3. CATEGORY BREAKDOWN
    # ---------------------------------------------------------
    elements.append(Paragraph("Category Breakdown", section_title_style))
    
    cat_data = [["Category", "Planned Budget", "Actual Spend", "Variance", "Used %"]]
    
    for cat in budget_summary.get('by_category', []):
        cat_planned = cat.get('planned_amount', 0)
        cat_actual = cat.get('actual_amount', 0)
        cat_var = cat_actual - cat_planned
        util = (cat_actual / cat_planned * 100) if cat_planned > 0 else 0
        
        cat_data.append([
            cat.get('category__name', 'Unknown'),
            f"{cat_planned:,.2f}",
            f"{cat_actual:,.2f}",
            f"{cat_var:,.2f}",
            f"{util:.1f}%"
        ])
        
    cat_table = Table(cat_data, colWidths=[150, 105, 105, 105, 65])
    cat_table.setStyle(TableStyle([
        # Header Row
        ('BACKGROUND', (0, 0), (-1, 0), colors.HexColor("#0f172a")),
        ('TEXTCOLOR', (0, 0), (-1, 0), colors.white),
        ('FONTNAME', (0, 0), (-1, 0), 'Helvetica-Bold'),
        ('FONTSIZE', (0, 0), (-1, 0), 10),
        ('ALIGN', (0, 0), (0, -1), 'LEFT'),
        ('ALIGN', (1, 0), (-1, -1), 'RIGHT'),
        ('BOTTOMPADDING', (0, 0), (-1, 0), 8),
        ('TOPPADDING', (0, 0), (-1, 0), 8),
        
        # Data Rows
        ('FONTNAME', (0, 1), (-1, -1), 'Helvetica'),
        ('FONTSIZE', (0, 1), (-1, -1), 10),
        ('TEXTCOLOR', (0, 1), (-1, -1), colors.HexColor("#334155")),
        ('BOTTOMPADDING', (0, 1), (-1, -1), 8),
        ('TOPPADDING', (0, 1), (-1, -1), 8),
        
        # Clean alternating rows with subtle lines
        ('ROWBACKGROUNDS', (0, 1), (-1, -1), [colors.white, colors.HexColor("#f8fafc")]),
        ('LINEBELOW', (0, 0), (-1, -1), 0.5, colors.HexColor("#e2e8f0")),
    ]))
    
    elements.append(cat_table)
    elements.append(Spacer(1, 30))
    
    # ---------------------------------------------------------
    # 4. RECENT TRANSACTIONS
    # ---------------------------------------------------------
    elements.append(Paragraph("Recent Approved Transactions", section_title_style))
    
    recent_txs = project.transactions.filter(status='approved').order_by('-transaction_date')[:10]
    
    if recent_txs.exists():
        tx_data = [["Date", "Description", "Category", "Amount (Rwf)"]]
        for tx in recent_txs:
            tx_data.append([
                tx.transaction_date.strftime("%b %d, %Y"),
                tx.description[:45] + ("..." if len(tx.description) > 45 else ""),
                tx.category.name if tx.category else "-",
                f"{tx.amount:,.2f}"
            ])
            
        tx_table = Table(tx_data, colWidths=[80, 230, 110, 110])
        tx_table.setStyle(TableStyle([
            # Header Row
            ('BACKGROUND', (0, 0), (-1, 0), colors.HexColor("#334155")),
            ('TEXTCOLOR', (0, 0), (-1, 0), colors.white),
            ('FONTNAME', (0, 0), (-1, 0), 'Helvetica-Bold'),
            ('FONTSIZE', (0, 0), (-1, 0), 10),
            ('ALIGN', (0, 0), (2, -1), 'LEFT'),
            ('ALIGN', (3, 0), (3, -1), 'RIGHT'),
            ('BOTTOMPADDING', (0, 0), (-1, 0), 8),
            ('TOPPADDING', (0, 0), (-1, 0), 8),
            
            # Data Rows
            ('FONTNAME', (0, 1), (-1, -1), 'Helvetica'),
            ('FONTSIZE', (0, 1), (-1, -1), 9),
            ('TEXTCOLOR', (0, 1), (-1, -1), colors.HexColor("#475569")),
            ('BOTTOMPADDING', (0, 1), (-1, -1), 6),
            ('TOPPADDING', (0, 1), (-1, -1), 6),
            
            ('ROWBACKGROUNDS', (0, 1), (-1, -1), [colors.white, colors.HexColor("#f8fafc")]),
            ('LINEBELOW', (0, 0), (-1, -1), 0.5, colors.HexColor("#e2e8f0")),
        ]))
        elements.append(tx_table)
    else:
        elements.append(Paragraph("No approved transactions found.", ParagraphStyle('i', parent=normal_style, fontName='Helvetica-Oblique', textColor=colors.gray)))
        
    append_report_footer(elements, styles, **meta)
    
    doc.build(elements)
    
    pdf_bytes = buffer.getvalue()
    buffer.close()
    
    return pdf_bytes

def _create_base_doc(buffer, project, subtitle, exported_by=None):
    doc = SimpleDocTemplate(buffer, pagesize=letter, rightMargin=40, leftMargin=40, topMargin=40, bottomMargin=40)
    styles = getSampleStyleSheet()
    header_title_style = ParagraphStyle('HeaderTitle', parent=styles['Heading1'], fontName='Helvetica-Bold', fontSize=24, textColor=colors.whitesmoke, alignment=TA_LEFT, spaceAfter=0)
    header_subtitle_style = ParagraphStyle('HeaderSubtitle', parent=styles['Normal'], fontName='Helvetica', fontSize=12, textColor=colors.HexColor("#cbd5e1"), alignment=TA_LEFT, spaceAfter=0)
    
    title_para = Paragraph("<b>BuildWise</b>", header_title_style)
    subtitle_para = Paragraph(subtitle, header_subtitle_style)
    project_info = Paragraph(
        build_header_project_info_html(project, exported_by=exported_by),
        header_project_info_style(),
    )
    
    header_table = Table([[[title_para, Spacer(1, 4), subtitle_para], project_info]], colWidths=[330, 200])
    header_table.setStyle(TableStyle([
        ('BACKGROUND', (0, 0), (-1, -1), colors.HexColor("#2563eb")),
        ('VALIGN', (0, 0), (-1, -1), 'MIDDLE'),
        ('TOPPADDING', (0, 0), (-1, -1), 20),
        ('BOTTOMPADDING', (0, 0), (-1, -1), 20),
        ('LEFTPADDING', (0, 0), (0, 0), 20),
        ('RIGHTPADDING', (1, 0), (1, 0), 20),
    ]))
    
    return doc, header_table, styles

def _get_table_style():
    return TableStyle([
        ('BACKGROUND', (0, 0), (-1, 0), colors.HexColor("#0f172a")),
        ('TEXTCOLOR', (0, 0), (-1, 0), colors.white),
        ('FONTNAME', (0, 0), (-1, 0), 'Helvetica-Bold'),
        ('FONTSIZE', (0, 0), (-1, 0), 10),
        ('BOTTOMPADDING', (0, 0), (-1, 0), 8),
        ('TOPPADDING', (0, 0), (-1, 0), 8),
        ('FONTNAME', (0, 1), (-1, -1), 'Helvetica'),
        ('FONTSIZE', (0, 1), (-1, -1), 9),
        ('TEXTCOLOR', (0, 1), (-1, -1), colors.HexColor("#334155")),
        ('BOTTOMPADDING', (0, 1), (-1, -1), 8),
        ('TOPPADDING', (0, 1), (-1, -1), 8),
        ('LEFTPADDING', (0, 0), (-1, -1), 6),
        ('RIGHTPADDING', (0, 0), (-1, -1), 6),
        ('VALIGN', (0, 0), (-1, -1), 'TOP'),
        ('ROWBACKGROUNDS', (0, 1), (-1, -1), [colors.white, colors.HexColor("#f8fafc")]),
        ('LINEBELOW', (0, 0), (-1, -1), 0.5, colors.HexColor("#e2e8f0")),
    ])


def _escape_cell_text(value):
    if value is None:
        return "—"
    return html.escape(str(value))


def _table_cell_styles(base_styles):
    return {
        'header': ParagraphStyle(
            'TableHeaderCell',
            parent=base_styles['Normal'],
            fontName='Helvetica-Bold',
            fontSize=10,
            textColor=colors.white,
            leading=12,
            wordWrap='LTR',
        ),
        'body': ParagraphStyle(
            'TableBodyCell',
            parent=base_styles['Normal'],
            fontName='Helvetica',
            fontSize=9,
            textColor=colors.HexColor("#334155"),
            leading=11,
            wordWrap='LTR',
        ),
    }


def _wrap_table_rows(headers, rows, styles):
    """Build table rows with Paragraph cells so long text wraps inside each column."""
    cell_styles = _table_cell_styles(styles)
    wrapped = [[Paragraph(_escape_cell_text(h), cell_styles['header']) for h in headers]]
    for row in rows:
        wrapped.append([Paragraph(_escape_cell_text(c), cell_styles['body']) for c in row])
    return wrapped


def _transaction_table_style(amount_col_index):
    """Table style for transaction reports with wrapped cells."""
    return TableStyle([
        ('BACKGROUND', (0, 0), (-1, 0), colors.HexColor("#0f172a")),
        ('TEXTCOLOR', (0, 0), (-1, 0), colors.whitesmoke),
        ('ALIGN', (0, 0), (-1, -1), 'LEFT'),
        ('ALIGN', (amount_col_index, 0), (amount_col_index, -1), 'RIGHT'),
        ('FONTNAME', (0, 0), (-1, 0), 'Helvetica-Bold'),
        ('FONTSIZE', (0, 0), (-1, 0), 10),
        ('BOTTOMPADDING', (0, 0), (-1, 0), 10),
        ('TOPPADDING', (0, 0), (-1, 0), 10),
        ('FONTNAME', (0, 1), (-1, -1), 'Helvetica'),
        ('FONTSIZE', (0, 1), (-1, -1), 9),
        ('BOTTOMPADDING', (0, 1), (-1, -1), 8),
        ('TOPPADDING', (0, 1), (-1, -1), 8),
        ('LEFTPADDING', (0, 0), (-1, -1), 6),
        ('RIGHTPADDING', (0, 0), (-1, -1), 6),
        ('VALIGN', (0, 0), (-1, -1), 'TOP'),
        ('GRID', (0, 0), (-1, -1), 1, colors.HexColor("#e2e8f0")),
        ('ROWBACKGROUNDS', (0, 1), (-1, -1), [colors.white, colors.HexColor("#f8fafc")]),
    ])

def generate_progress_report_pdf(project, exported_by=None, verification_token=None, verify_url=None):
    meta = _meta_kwargs(exported_by, verification_token, verify_url)
    buffer = io.BytesIO()
    doc, header, styles = _create_base_doc(buffer, project, "Monthly Progress Report", exported_by=exported_by)
    elements = [header, Spacer(1, 30)]
    
    section_title_style = ParagraphStyle('ST', parent=styles['Heading2'], fontName='Helvetica-Bold', fontSize=16, textColor=colors.HexColor("#1e293b"), spaceAfter=10)
    elements.append(Paragraph("Project Status Overview", section_title_style))
    
    tasks = project.tasks.all()
    total = tasks.count()
    completed = tasks.filter(status='completed').count()
    pending = tasks.filter(status='pending').count()
    in_prog = tasks.filter(status='in_progress').count()
    
    summary_data = [["TOTAL TASKS", "COMPLETED", "IN PROGRESS", "PENDING"],
                    [str(total), str(completed), str(in_prog), str(pending)]]
    summary_table = Table(summary_data, colWidths=[132]*4)
    summary_table.setStyle(TableStyle([
        ('BACKGROUND', (0, 0), (-1, 0), colors.HexColor("#f1f5f9")),
        ('TEXTCOLOR', (0, 0), (-1, 0), colors.HexColor("#64748b")),
        ('FONTNAME', (0, 0), (-1, 0), 'Helvetica-Bold'),
        ('FONTSIZE', (0, 0), (-1, 0), 9),
        ('ALIGN', (0, 0), (-1, -1), 'CENTER'),
        ('VALIGN', (0, 0), (-1, -1), 'MIDDLE'),
        ('FONTNAME', (0, 1), (-1, 1), 'Helvetica-Bold'),
        ('FONTSIZE', (0, 1), (-1, 1), 12),
        ('TEXTCOLOR', (0, 1), (-1, 1), colors.HexColor("#0f172a")),
        ('BOTTOMPADDING', (0, 1), (-1, 1), 15),
        ('TOPPADDING', (0, 0), (-1, 1), 10),
    ]))
    elements.extend([summary_table, Spacer(1, 30)])
    
    elements.append(Paragraph("Recent Tasks", section_title_style))
    task_data = [["Task", "Status", "Priority", "Date"]]
    for t in tasks.order_by('-updated_at')[:15]:
        task_data.append([t.title[:40], t.get_status_display(), t.get_priority_display(), t.date.strftime("%b %d, %Y") if t.date else "-"])
    
    if len(task_data) > 1:
        t_table = Table(task_data, colWidths=[230, 100, 100, 100])
        t_table.setStyle(_get_table_style())
        elements.append(t_table)
    else:
        elements.append(Paragraph("No tasks available.", styles['Normal']))

    append_report_footer(elements, styles, **meta)
    doc.build(elements)
    pdf_bytes = buffer.getvalue()
    buffer.close()
    return pdf_bytes

def generate_hr_report_pdf(project, exported_by=None, verification_token=None, verify_url=None):
    meta = _meta_kwargs(exported_by, verification_token, verify_url)
    buffer = io.BytesIO()
    doc, header, styles = _create_base_doc(buffer, project, "Resource Utilization (HR) Report", exported_by=exported_by)
    elements = [header, Spacer(1, 30)]
    
    section_title_style = ParagraphStyle('ST', parent=styles['Heading2'], fontName='Helvetica-Bold', fontSize=16, textColor=colors.HexColor("#1e293b"), spaceAfter=10)
    
    workers = project.workers.all()
    active = workers.filter(is_active=True).count()
    elements.append(Paragraph("Workforce Summary", section_title_style))
    elements.append(Paragraph(f"<b>Total Registered Workers:</b> {workers.count()}<br/><b>Active Workers:</b> {active}", styles['Normal']))
    elements.append(Spacer(1, 20))
    
    elements.append(Paragraph("Active Workers Roster", section_title_style))
    w_data = [["Name", "Role", "Phone", "Daily Rate"]]
    for w in workers.filter(is_active=True)[:20]:
        w_data.append([f"{w.first_name} {w.last_name}", w.get_role_display(), w.phone_number or "-", f"Rwf {w.daily_rate:,.2f}"])
        
    if len(w_data) > 1:
        w_table = Table(w_data, colWidths=[150, 130, 120, 130])
        w_table.setStyle(_get_table_style())
        elements.append(w_table)
    else:
        elements.append(Paragraph("No active workers.", styles['Normal']))

    append_report_footer(elements, styles, **meta)
    doc.build(elements)
    pdf_bytes = buffer.getvalue()
    buffer.close()
    return pdf_bytes

def generate_procurement_report_pdf(project, exported_by=None, verification_token=None, verify_url=None):
    from procurement.models import PurchaseOrder, MaterialRequest

    meta = _meta_kwargs(exported_by, verification_token, verify_url)
    buffer = io.BytesIO()
    doc, header, styles = _create_base_doc(buffer, project, "Material Inventory Log", exported_by=exported_by)
    elements = [header, Spacer(1, 30)]

    section_title_style = ParagraphStyle(
        'ST',
        parent=styles['Heading2'],
        fontName='Helvetica-Bold',
        fontSize=16,
        textColor=colors.HexColor("#1e293b"),
        spaceAfter=10,
        spaceBefore=10,
    )

    # Site inventory (per-project stock)
    elements.append(Paragraph("Site Inventory", section_title_style))
    site_items = project.site_inventory.select_related('material').order_by('material__name')
    i_data = [["Material", "Unit", "Qty on Site", "Unit Price", "Est. Value"]]
    for row in site_items[:25]:
        mat = row.material
        qty = float(row.current_stock or 0)
        price = float(mat.unit_price or 0)
        val = qty * price
        i_data.append([
            (mat.name or "")[:32],
            (mat.unit or "")[:12],
            f"{qty:,.2f}",
            f"Rwf {price:,.2f}",
            f"Rwf {val:,.2f}",
        ])

    if len(i_data) > 1:
        i_table = Table(i_data, colWidths=[150, 70, 90, 100, 110])
        i_table.setStyle(_get_table_style())
        elements.append(i_table)
    else:
        elements.append(Paragraph("No site inventory recorded for this project.", styles['Normal']))

    elements.append(Spacer(1, 24))

    # Purchase orders
    elements.append(Paragraph("Purchase Orders", section_title_style))
    orders = PurchaseOrder.objects.filter(project=project).select_related(
        'supplier', 'material'
    ).order_by('-order_date')[:20]
    po_data = [["PO #", "Supplier", "Material", "Qty", "Amount", "Status"]]
    po_total = 0.0
    for po in orders:
        amt = float(po.total_amount or 0)
        po_total += amt
        po_data.append([
            (po.po_number or "")[:14],
            (po.supplier.name if po.supplier else "—")[:18],
            (po.material.name if po.material else "—")[:18],
            f"{float(po.quantity or 0):,.1f}",
            f"Rwf {amt:,.2f}",
            po.get_status_display(),
        ])

    if len(po_data) > 1:
        po_table = Table(po_data, colWidths=[70, 95, 95, 55, 95, 70])
        po_table.setStyle(_get_table_style())
        elements.append(po_table)
        elements.append(Spacer(1, 8))
        elements.append(Paragraph(
            f"<b>Total PO value (listed):</b> Rwf {po_total:,.2f}",
            styles['Normal'],
        ))
    else:
        elements.append(Paragraph("No purchase orders for this project.", styles['Normal']))

    elements.append(Spacer(1, 24))

    # Material requests
    elements.append(Paragraph("Material Requests", section_title_style))
    requests = MaterialRequest.objects.filter(project=project).select_related(
        'material', 'requested_by'
    ).order_by('-created_at')[:15]
    mr_data = [["Material", "Qty", "Status", "Requested By", "Date"]]
    for req in requests:
        mr_data.append([
            (req.material.name if req.material else "—")[:24],
            f"{float(req.quantity_requested or 0):,.1f}",
            req.get_status_display(),
            (req.requested_by.full_name or req.requested_by.username)[:20] if req.requested_by else "—",
            req.created_at.strftime("%b %d, %Y") if req.created_at else "—",
        ])

    if len(mr_data) > 1:
        mr_table = Table(mr_data, colWidths=[130, 60, 90, 120, 90])
        mr_table.setStyle(_get_table_style())
        elements.append(mr_table)
    else:
        elements.append(Paragraph("No material requests for this project.", styles['Normal']))

    append_report_footer(elements, styles, **meta)
    doc.build(elements)
    pdf_bytes = buffer.getvalue()
    buffer.close()
    return pdf_bytes

def generate_daily_transactions_report_pdf(project, exported_by=None, verification_token=None, verify_url=None):
    from django.utils import timezone
    from .models import Transaction

    meta = _meta_kwargs(exported_by, verification_token, verify_url)
    buffer = io.BytesIO()
    doc = SimpleDocTemplate(
        buffer, pagesize=letter, rightMargin=40, leftMargin=40, topMargin=40, bottomMargin=40
    )
    elements = []
    
    styles = getSampleStyleSheet()
    header_title_style = ParagraphStyle(
        'HeaderTitle', parent=styles['Heading1'], fontName='Helvetica-Bold', fontSize=24, textColor=colors.whitesmoke, alignment=TA_LEFT, spaceAfter=0
    )
    header_subtitle_style = ParagraphStyle(
        'HeaderSubtitle', parent=styles['Normal'], fontName='Helvetica', fontSize=12, textColor=colors.HexColor("#cbd5e1"), alignment=TA_LEFT, spaceAfter=0
    )
    section_title_style = ParagraphStyle(
        'SectionTitle', parent=styles['Heading2'], fontName='Helvetica-Bold', fontSize=16, textColor=colors.HexColor("#1e293b"), spaceAfter=10, spaceBefore=20
    )
    
    today = timezone.now().date()
    
    title_para = Paragraph("<b>BuildWise</b>", header_title_style)
    subtitle_para = Paragraph("Daily Transactions Report", header_subtitle_style)
    project_info = Paragraph(
        build_header_project_info_html(
            project,
            exported_by=exported_by,
            extra_lines=[f"<b>Report date:</b> {today.strftime('%B %d, %Y')}"],
        ),
        header_project_info_style(),
    )
    
    header_data = [[[title_para, Spacer(1, 4), subtitle_para], project_info]]
    header_table = Table(header_data, colWidths=[330, 200])
    header_table.setStyle(TableStyle([
        ('BACKGROUND', (0, 0), (-1, -1), colors.HexColor("#4f46e5")),
        ('VALIGN', (0, 0), (-1, -1), 'MIDDLE'),
        ('TOPPADDING', (0, 0), (-1, -1), 20),
        ('BOTTOMPADDING', (0, 0), (-1, -1), 20),
        ('LEFTPADDING', (0, 0), (0, 0), 20),
        ('RIGHTPADDING', (1, 0), (1, 0), 20),
    ]))
    
    elements.append(header_table)
    elements.append(Spacer(1, 30))
    
    elements.append(Paragraph("Transactions Logged Today", section_title_style))
    
    today_transactions = Transaction.objects.filter(project=project, transaction_date=today).order_by('-id')
    
    if not today_transactions.exists():
        elements.append(Paragraph("No transactions were logged today.", styles['Normal']))
    else:
        tx_data = [["ID", "Description", "Category", "Status", "Amount (Rwf)"]]
        total_amount = 0
        
        for tx in today_transactions:
            cat_name = tx.category.name if tx.category else "N/A"
            tx_data.append([
                f"TRX-{tx.id}",
                tx.description or "—",
                cat_name,
                tx.status.capitalize(),
                f"{tx.amount:,.2f}"
            ])
            if tx.status == 'approved':
                total_amount += tx.amount

        wrapped_tx = _wrap_table_rows(tx_data[0], tx_data[1:], styles)
        tx_table = Table(wrapped_tx, colWidths=[60, 190, 100, 80, 100], repeatRows=1)
        tx_table.setStyle(_transaction_table_style(4))
        
        elements.append(tx_table)
        elements.append(Spacer(1, 20))
        
        summary_style = ParagraphStyle(
            'TxSummary', parent=styles['Normal'], fontName='Helvetica-Bold', fontSize=12, alignment=TA_RIGHT
        )
        elements.append(Paragraph(f"Total Approved Spending Today: Rwf {total_amount:,.2f}", summary_style))

    append_report_footer(elements, styles, **meta)
    doc.build(elements)
    pdf_bytes = buffer.getvalue()
    buffer.close()
    return pdf_bytes

def generate_custom_transactions_report_pdf(project, start_date_str, end_date_str, exported_by=None, verification_token=None, verify_url=None):
    from django.utils import timezone
    from .models import Transaction
    from datetime import datetime

    meta = _meta_kwargs(exported_by, verification_token, verify_url)
    buffer = io.BytesIO()
    doc = SimpleDocTemplate(
        buffer, pagesize=letter, rightMargin=40, leftMargin=40, topMargin=40, bottomMargin=40
    )
    elements = []
    
    styles = getSampleStyleSheet()
    header_title_style = ParagraphStyle(
        'HeaderTitle', parent=styles['Heading1'], fontName='Helvetica-Bold', fontSize=24, textColor=colors.whitesmoke, alignment=TA_LEFT, spaceAfter=0
    )
    header_subtitle_style = ParagraphStyle(
        'HeaderSubtitle', parent=styles['Normal'], fontName='Helvetica', fontSize=12, textColor=colors.HexColor("#cbd5e1"), alignment=TA_LEFT, spaceAfter=0
    )
    section_title_style = ParagraphStyle(
        'SectionTitle', parent=styles['Heading2'], fontName='Helvetica-Bold', fontSize=16, textColor=colors.HexColor("#1e293b"), spaceAfter=10, spaceBefore=20
    )
    
    try:
        start_dt = datetime.strptime(start_date_str, '%Y-%m-%d').date()
        end_dt = datetime.strptime(end_date_str, '%Y-%m-%d').date()
        date_range_str = f"{start_dt.strftime('%b %d, %Y')} - {end_dt.strftime('%b %d, %Y')}"
    except:
        start_dt = timezone.now().date()
        end_dt = timezone.now().date()
        date_range_str = "Invalid Date Range"

    title_para = Paragraph("<b>BuildWise</b>", header_title_style)
    subtitle_para = Paragraph("Custom Transactions Report", header_subtitle_style)
    project_info = Paragraph(
        build_header_project_info_html(
            project,
            exported_by=exported_by,
            extra_lines=[f"<b>Range:</b> {date_range_str}"],
        ),
        header_project_info_style(),
    )
    
    header_data = [[[title_para, Spacer(1, 4), subtitle_para], project_info]]
    header_table = Table(header_data, colWidths=[330, 200])
    header_table.setStyle(TableStyle([
        ('BACKGROUND', (0, 0), (-1, -1), colors.HexColor("#334155")), # Slate 700
        ('VALIGN', (0, 0), (-1, -1), 'MIDDLE'),
        ('TOPPADDING', (0, 0), (-1, -1), 20),
        ('BOTTOMPADDING', (0, 0), (-1, -1), 20),
        ('LEFTPADDING', (0, 0), (0, 0), 20),
        ('RIGHTPADDING', (1, 0), (1, 0), 20),
    ]))
    
    elements.append(header_table)
    elements.append(Spacer(1, 30))
    
    elements.append(Paragraph(f"Transactions from {date_range_str}", section_title_style))
    
    transactions = Transaction.objects.filter(project=project, transaction_date__gte=start_dt, transaction_date__lte=end_dt).order_by('transaction_date', '-id')
    
    if not transactions.exists():
        elements.append(Paragraph("No transactions were logged in this period.", styles['Normal']))
    else:
        tx_data = [["Date", "Description", "Category", "Status", "Amount (Rwf)"]]
        total_amount = 0
        
        for tx in transactions:
            cat_name = tx.category.name if tx.category else "N/A"
            tx_data.append([
                tx.transaction_date.strftime("%b %d"),
                tx.description or "—",
                cat_name,
                tx.status.capitalize(),
                f"{tx.amount:,.2f}"
            ])
            if tx.status == 'approved':
                total_amount += tx.amount

        wrapped_tx = _wrap_table_rows(tx_data[0], tx_data[1:], styles)
        tx_table = Table(wrapped_tx, colWidths=[65, 175, 90, 70, 90], repeatRows=1)
        tx_table.setStyle(_transaction_table_style(4))
        
        elements.append(tx_table)
        elements.append(Spacer(1, 20))
        
        summary_style = ParagraphStyle(
            'TxSummary', parent=styles['Normal'], fontName='Helvetica-Bold', fontSize=12, alignment=TA_RIGHT
        )
        elements.append(Paragraph(f"Total Approved Spending in Period: Rwf {total_amount:,.2f}", summary_style))

    append_report_footer(elements, styles, **meta)
    doc.build(elements)
    pdf_bytes = buffer.getvalue()
    buffer.close()
    return pdf_bytes


def _timeline_querysets(project):
    from .models import PhaseTask, Milestone, ProjectPhase

    phase_tasks = (
        PhaseTask.objects.filter(project=project)
        .select_related('project_phase')
        .prefetch_related('assigned_to')
        .order_by('start_date', 'id')
    )
    milestones = Milestone.objects.filter(project=project).order_by('date')
    phases = ProjectPhase.objects.filter(project=project).order_by('order', 'id')
    return phase_tasks, milestones, phases


def generate_timeline_report_pdf(project, exported_by=None, verification_token=None, verify_url=None):
    meta = _meta_kwargs(exported_by, verification_token, verify_url)
    phase_tasks, milestones, phases = _timeline_querysets(project)

    buffer = io.BytesIO()
    doc, header, styles = _create_base_doc(buffer, project, "Project Timeline Report", exported_by=exported_by)
    elements = [header, Spacer(1, 24)]

    section_title_style = ParagraphStyle(
        'ST',
        parent=styles['Heading2'],
        fontName='Helvetica-Bold',
        fontSize=16,
        textColor=colors.HexColor("#1e293b"),
        spaceAfter=10,
        spaceBefore=10,
    )

    total_tasks = phase_tasks.count()
    completed = phase_tasks.filter(status='completed').count()
    on_track = phase_tasks.filter(status='on-track').count()
    pending = phase_tasks.filter(status='pending').count()
    delayed = phase_tasks.filter(status='delayed').count()

    summary_data = [
        ["TOTAL TASKS", "COMPLETED", "IN PROGRESS", "PENDING", "DELAYED"],
        [str(total_tasks), str(completed), str(on_track), str(pending), str(delayed)],
    ]
    summary_table = Table(summary_data, colWidths=[106, 106, 106, 106, 106])
    summary_table.setStyle(TableStyle([
        ('BACKGROUND', (0, 0), (-1, 0), colors.HexColor("#f1f5f9")),
        ('TEXTCOLOR', (0, 0), (-1, 0), colors.HexColor("#64748b")),
        ('FONTNAME', (0, 0), (-1, 0), 'Helvetica-Bold'),
        ('FONTSIZE', (0, 0), (-1, 0), 8),
        ('ALIGN', (0, 0), (-1, -1), 'CENTER'),
        ('FONTNAME', (0, 1), (-1, 1), 'Helvetica-Bold'),
        ('FONTSIZE', (0, 1), (-1, 1), 11),
        ('TOPPADDING', (0, 0), (-1, -1), 8),
        ('BOTTOMPADDING', (0, 0), (-1, -1), 8),
    ]))
    elements.append(Paragraph("Timeline Overview", section_title_style))
    elements.append(summary_table)
    elements.append(Spacer(1, 20))

    elements.append(Paragraph("Scheduled Tasks", section_title_style))
    task_data = [["Phase", "Task", "Start", "End", "Duration", "Assigned To", "Status", "Progress"]]
    for task in phase_tasks:
        assignees = ", ".join(
            f"{w.first_name} {w.last_name}".strip() or "Worker"
            for w in task.assigned_to.all()
        ) if task.assigned_to.exists() else "Unassigned"
        phase_label = task.project_phase.name if task.project_phase_id else (task.phase or "—")
        task_data.append([
            (phase_label or "—")[:22],
            (task.task_name or "—")[:28],
            task.start_date.strftime("%b %d, %Y") if task.start_date else "—",
            task.end_date.strftime("%b %d, %Y") if task.end_date else "—",
            str(task.duration_working_days or "—"),
            assignees[:24],
            task.get_status_display(),
            f"{task.progress or 0}%",
        ])

    if len(task_data) > 1:
        t_table = Table(task_data, colWidths=[72, 95, 62, 62, 48, 78, 58, 45])
        t_table.setStyle(_get_table_style())
        elements.append(t_table)
    else:
        elements.append(Paragraph("No scheduled tasks on this timeline.", styles['Normal']))

    elements.append(Spacer(1, 20))
    elements.append(Paragraph("Milestones", section_title_style))
    ms_data = [["Milestone", "Target Date", "Status"]]
    for ms in milestones:
        ms_data.append([
            (ms.name or "—")[:40],
            ms.date.strftime("%b %d, %Y") if ms.date else "—",
            ms.get_status_display() if hasattr(ms, 'get_status_display') else (ms.status or "—"),
        ])

    if len(ms_data) > 1:
        ms_table = Table(ms_data, colWidths=[280, 120, 130])
        ms_table.setStyle(_get_table_style())
        elements.append(ms_table)
    else:
        elements.append(Paragraph("No milestones defined.", styles['Normal']))

    if phases.exists():
        elements.append(Spacer(1, 20))
        elements.append(Paragraph("Construction Phases", section_title_style))
        ph_data = [["Phase", "Start", "End", "Description"]]
        for ph in phases:
            ph_data.append([
                (ph.name or "—")[:30],
                ph.start_date.strftime("%b %d, %Y") if ph.start_date else "—",
                ph.end_date.strftime("%b %d, %Y") if ph.end_date else "—",
                (ph.description or "—")[:40],
            ])
        ph_table = Table(ph_data, colWidths=[120, 90, 90, 230])
        ph_table.setStyle(_get_table_style())
        elements.append(ph_table)

    append_report_footer(elements, styles, **meta)
    doc.build(elements)
    pdf_bytes = buffer.getvalue()
    buffer.close()
    return pdf_bytes


def generate_timeline_excel(project):
    from openpyxl import Workbook
    from openpyxl.styles import Font, PatternFill

    phase_tasks, milestones, phases = _timeline_querysets(project)
    wb = Workbook()

    ws_tasks = wb.active
    ws_tasks.title = "Timeline Tasks"
    task_headers = ["Phase", "Task Name", "Start Date", "End Date", "Duration (Days)", "Assigned To", "Status", "Progress %"]
    ws_tasks.append(task_headers)
    for cell in ws_tasks[1]:
        cell.font = Font(bold=True, color="FFFFFF")
        cell.fill = PatternFill("solid", fgColor="2563EB")

    for task in phase_tasks:
        assignees = ", ".join(
            f"{w.first_name} {w.last_name}".strip() or "Worker"
            for w in task.assigned_to.all()
        ) if task.assigned_to.exists() else "Unassigned"
        phase_label = task.project_phase.name if task.project_phase_id else (task.phase or "")
        ws_tasks.append([
            phase_label,
            task.task_name,
            task.start_date.isoformat() if task.start_date else "",
            task.end_date.isoformat() if task.end_date else "",
            task.duration_working_days,
            assignees,
            task.status,
            float(task.progress or 0),
        ])

    ws_ms = wb.create_sheet("Milestones")
    ms_headers = ["Milestone", "Target Date", "Status"]
    ws_ms.append(ms_headers)
    for cell in ws_ms[1]:
        cell.font = Font(bold=True, color="FFFFFF")
        cell.fill = PatternFill("solid", fgColor="2563EB")
    for ms in milestones:
        ws_ms.append([
            ms.name,
            ms.date.isoformat() if ms.date else "",
            ms.status,
        ])

    ws_ph = wb.create_sheet("Phases")
    ph_headers = ["Phase", "Start Date", "End Date", "Description"]
    ws_ph.append(ph_headers)
    for cell in ws_ph[1]:
        cell.font = Font(bold=True, color="FFFFFF")
        cell.fill = PatternFill("solid", fgColor="2563EB")
    for ph in phases:
        ws_ph.append([
            ph.name,
            ph.start_date.isoformat() if ph.start_date else "",
            ph.end_date.isoformat() if ph.end_date else "",
            ph.description or "",
        ])

    buffer = io.BytesIO()
    wb.save(buffer)
    buffer.seek(0)
    return buffer.getvalue()


SITE_INVENTORY_REPORT_DAYS = 7


def _site_inventory_stock_status(stock, minimum_stock):
    stock = float(stock or 0)
    minimum = float(minimum_stock or 0)
    if stock <= 0:
        return 'Out of stock'
    if minimum > 0 and stock <= minimum:
        return 'Low stock'
    if minimum <= 0 and 0 < stock < 10:
        return 'Low stock'
    return 'In stock'


def _site_inventory_querysets(project, *, usage_date=None):
    from datetime import datetime, timedelta
    from django.utils import timezone
    from procurement.models import SiteInventory, MaterialAllocation, MaterialRequest

    items = (
        SiteInventory.objects.filter(project=project)
        .select_related('material')
        .order_by('material__name')
    )
    allocations_qs = (
        MaterialAllocation.objects.filter(site_inventory__project=project)
        .select_related(
            'site_inventory__material',
            'task',
            'allocated_by',
        )
        .order_by('-date_allocated')
    )
    cutoff = timezone.now() - timedelta(days=SITE_INVENTORY_REPORT_DAYS)
    recent_allocations = allocations_qs.filter(date_allocated__gte=cutoff)
    pending_requests = (
        MaterialRequest.objects.filter(project=project, status='pending')
        .select_related('material')
        .order_by('-created_at')
    )

    day_allocations = None
    if usage_date:
        day_allocations = allocations_qs.filter(date_allocated__date=usage_date)

    return {
        'items': items,
        'recent_allocations': recent_allocations,
        'pending_requests': pending_requests,
        'day_allocations': day_allocations,
    }


def _section_title_style(styles):
    return ParagraphStyle(
        'SiteInvSection',
        parent=styles['Heading2'],
        fontName='Helvetica-Bold',
        fontSize=16,
        textColor=colors.HexColor("#1e293b"),
        spaceAfter=10,
        spaceBefore=10,
    )


def generate_site_inventory_snapshot_pdf(project, exported_by=None, verification_token=None, verify_url=None):
    meta = _meta_kwargs(exported_by, verification_token, verify_url)
    data = _site_inventory_querysets(project)
    items = data['items']
    recent = data['recent_allocations']
    pending = data['pending_requests']

    low_count = out_count = 0
    for item in items:
        status = _site_inventory_stock_status(item.current_stock, item.material.minimum_stock)
        if status == 'Out of stock':
            out_count += 1
        elif status == 'Low stock':
            low_count += 1

    units_recent = sum(float(a.quantity) for a in recent)

    buffer = io.BytesIO()
    doc, header, styles = _create_base_doc(
        buffer, project, "Site Inventory — Full Report", exported_by=exported_by,
    )
    section_style = _section_title_style(styles)
    elements = [header, Spacer(1, 24)]

    summary_data = [
        ["TOTAL ITEMS", "LOW STOCK", "OUT OF STOCK", "PENDING REQUESTS", f"ALLOCATIONS ({SITE_INVENTORY_REPORT_DAYS}D)", "UNITS ISSUED (7D)"],
        [
            str(items.count()),
            str(low_count),
            str(out_count),
            str(pending.count()),
            str(recent.count()),
            f"{units_recent:,.2f}",
        ],
    ]
    summary_table = Table(summary_data, colWidths=[88, 88, 88, 88, 88, 88])
    summary_table.setStyle(TableStyle([
        ('BACKGROUND', (0, 0), (-1, 0), colors.HexColor("#f1f5f9")),
        ('TEXTCOLOR', (0, 0), (-1, 0), colors.HexColor("#64748b")),
        ('FONTNAME', (0, 0), (-1, 0), 'Helvetica-Bold'),
        ('FONTSIZE', (0, 0), (-1, 0), 8),
        ('ALIGN', (0, 0), (-1, -1), 'CENTER'),
        ('FONTNAME', (0, 1), (-1, 1), 'Helvetica-Bold'),
        ('FONTSIZE', (0, 1), (-1, 1), 11),
        ('TOPPADDING', (0, 0), (-1, -1), 8),
        ('BOTTOMPADDING', (0, 0), (-1, -1), 8),
    ]))
    elements.append(Paragraph("Inventory Overview", section_style))
    elements.append(summary_table)
    elements.append(Spacer(1, 20))

    elements.append(Paragraph("Current Stock Levels", section_style))
    inv_data = [["Material", "Unit", "On Hand", "Minimum", "Status", "Last Updated"]]
    for item in items:
        inv_data.append([
            item.material.name[:32],
            item.material.unit[:10],
            f"{float(item.current_stock):,.2f}",
            f"{float(item.material.minimum_stock):,.2f}",
            _site_inventory_stock_status(item.current_stock, item.material.minimum_stock),
            timezone_format(item.updated_at),
        ])
    if len(inv_data) > 1:
        inv_table = Table(inv_data, colWidths=[130, 45, 65, 65, 75, 90])
        inv_table.setStyle(_get_table_style())
        elements.append(inv_table)
    else:
        elements.append(Paragraph("No inventory items recorded for this site.", styles['Normal']))

    elements.append(Spacer(1, 20))
    elements.append(Paragraph(
        f"Recent Allocations (Last {SITE_INVENTORY_REPORT_DAYS} Days)",
        section_style,
    ))
    alloc_data = [["Date", "Material", "Qty", "Unit", "Task", "Allocated By", "Notes"]]
    for alloc in recent[:50]:
        material = alloc.site_inventory.material
        by_name = alloc.allocated_by.full_name or alloc.allocated_by.username if alloc.allocated_by_id else "—"
        alloc_data.append([
            alloc.date_allocated.strftime("%b %d, %Y %I:%M %p"),
            material.name[:24],
            f"{float(alloc.quantity):,.2f}",
            material.unit[:8],
            (getattr(alloc.task, 'title', None) or getattr(alloc.task, 'task_name', None) or "—")[:22],
            by_name[:18],
            (alloc.notes or "")[:28],
        ])
    if len(alloc_data) > 1:
        a_table = Table(alloc_data, colWidths=[78, 95, 42, 38, 85, 72, 80])
        a_table.setStyle(_get_table_style())
        elements.append(a_table)
    else:
        elements.append(Paragraph("No allocations in the reporting window.", styles['Normal']))

    if pending.exists():
        elements.append(Spacer(1, 20))
        elements.append(Paragraph("Pending Material Requests", section_style))
        req_data = [["Material", "Quantity", "Unit", "Requested", "Status"]]
        for req in pending[:30]:
            req_data.append([
                req.material.name[:32],
                f"{float(req.quantity_requested):,.2f}",
                req.material.unit[:10],
                req.created_at.strftime("%b %d, %Y"),
                req.get_status_display(),
            ])
        req_table = Table(req_data, colWidths=[150, 70, 50, 80, 100])
        req_table.setStyle(_get_table_style())
        elements.append(req_table)

    append_report_footer(elements, styles, **meta)
    doc.build(elements)
    pdf_bytes = buffer.getvalue()
    buffer.close()
    return pdf_bytes


def timezone_format(dt):
    from django.utils import timezone as dj_tz
    if dt is None:
        return "—"
    if dj_tz.is_aware(dt):
        dt = dj_tz.localtime(dt)
    return dt.strftime("%b %d, %Y %I:%M %p")


def generate_site_inventory_usage_pdf(
    project,
    usage_date,
    *,
    exported_by=None,
    verification_token=None,
    verify_url=None,
):
    from datetime import datetime

    meta = _meta_kwargs(exported_by, verification_token, verify_url)
    if isinstance(usage_date, str):
        usage_date = datetime.strptime(usage_date, '%Y-%m-%d').date()

    data = _site_inventory_querysets(project, usage_date=usage_date)
    day_allocations = list(data['day_allocations'] or [])

    usage_map = {}
    for alloc in day_allocations:
        material = alloc.site_inventory.material
        key = material.name
        entry = usage_map.get(key)
        if not entry:
            entry = {'material_name': material.name, 'unit': material.unit, 'total': 0.0, 'count': 0}
            usage_map[key] = entry
        entry['total'] += float(alloc.quantity)
        entry['count'] += 1

    period_label = usage_date.strftime("%A, %B %d, %Y")
    units_total = sum(float(a.quantity) for a in day_allocations)
    materials_count = len(usage_map)
    tasks_count = len({a.task_id for a in day_allocations})

    buffer = io.BytesIO()
    extra = [f"<b>Usage date:</b> {period_label}"]
    doc = SimpleDocTemplate(buffer, pagesize=letter, rightMargin=40, leftMargin=40, topMargin=40, bottomMargin=40)
    styles = getSampleStyleSheet()
    header_title_style = ParagraphStyle(
        'HeaderTitle', parent=styles['Heading1'], fontName='Helvetica-Bold',
        fontSize=24, textColor=colors.whitesmoke, alignment=TA_LEFT, spaceAfter=0,
    )
    header_subtitle_style = ParagraphStyle(
        'HeaderSubtitle', parent=styles['Normal'], fontName='Helvetica',
        fontSize=12, textColor=colors.HexColor("#cbd5e1"), alignment=TA_LEFT, spaceAfter=0,
    )
    title_para = Paragraph("<b>BuildWise</b>", header_title_style)
    subtitle_para = Paragraph("Site Inventory — Daily Usage Report", header_subtitle_style)
    project_info = Paragraph(
        build_header_project_info_html(project, exported_by=exported_by, extra_lines=extra),
        header_project_info_style(),
    )
    header_table = Table([[[title_para, Spacer(1, 4), subtitle_para], project_info]], colWidths=[330, 200])
    header_table.setStyle(TableStyle([
        ('BACKGROUND', (0, 0), (-1, -1), colors.HexColor("#2563eb")),
        ('VALIGN', (0, 0), (-1, -1), 'MIDDLE'),
        ('TOPPADDING', (0, 0), (-1, -1), 20),
        ('BOTTOMPADDING', (0, 0), (-1, -1), 20),
        ('LEFTPADDING', (0, 0), (0, 0), 20),
        ('RIGHTPADDING', (1, 0), (1, 0), 20),
    ]))

    section_style = _section_title_style(styles)
    elements = [header_table, Spacer(1, 24)]

    summary_data = [
        ["ALLOCATIONS", "UNITS ISSUED", "MATERIALS", "TASKS SUPPLIED"],
        [str(len(day_allocations)), f"{units_total:,.2f}", str(materials_count), str(tasks_count)],
    ]
    summary_table = Table(summary_data, colWidths=[132, 132, 132, 132])
    summary_table.setStyle(TableStyle([
        ('BACKGROUND', (0, 0), (-1, 0), colors.HexColor("#f1f5f9")),
        ('TEXTCOLOR', (0, 0), (-1, 0), colors.HexColor("#64748b")),
        ('FONTNAME', (0, 0), (-1, 0), 'Helvetica-Bold'),
        ('FONTSIZE', (0, 0), (-1, 0), 8),
        ('ALIGN', (0, 0), (-1, -1), 'CENTER'),
        ('FONTNAME', (0, 1), (-1, 1), 'Helvetica-Bold'),
        ('FONTSIZE', (0, 1), (-1, 1), 11),
        ('TOPPADDING', (0, 0), (-1, -1), 8),
        ('BOTTOMPADDING', (0, 0), (-1, -1), 8),
    ]))
    elements.append(Paragraph(f"Usage Summary — {period_label}", section_style))
    elements.append(summary_table)
    elements.append(Spacer(1, 20))

    elements.append(Paragraph("Usage by Material", section_style))
    usage_data = [["Material", "Unit", "Total Quantity", "Allocations"]]
    for entry in sorted(usage_map.values(), key=lambda x: x['total'], reverse=True):
        usage_data.append([
            entry['material_name'][:36],
            entry['unit'][:10],
            f"{entry['total']:,.2f}",
            str(entry['count']),
        ])
    if len(usage_data) > 1:
        u_table = Table(usage_data, colWidths=[200, 60, 100, 90])
        u_table.setStyle(_get_table_style())
        elements.append(u_table)
    else:
        elements.append(Paragraph("No material usage recorded for this date.", styles['Normal']))

    elements.append(Spacer(1, 20))
    elements.append(Paragraph("Allocation Detail", section_style))
    detail_data = [["Time", "Material", "Qty", "Unit", "Task", "Allocated By", "Notes"]]
    for alloc in day_allocations:
        material = alloc.site_inventory.material
        by_name = alloc.allocated_by.full_name or alloc.allocated_by.username if alloc.allocated_by_id else "—"
        detail_data.append([
            alloc.date_allocated.strftime("%I:%M %p"),
            material.name[:24],
            f"{float(alloc.quantity):,.2f}",
            material.unit[:8],
            (getattr(alloc.task, 'title', None) or getattr(alloc.task, 'task_name', None) or "—")[:22],
            by_name[:18],
            (alloc.notes or "")[:28],
        ])
    if len(detail_data) > 1:
        d_table = Table(detail_data, colWidths=[52, 95, 42, 38, 85, 72, 80])
        d_table.setStyle(_get_table_style())
        elements.append(d_table)

    append_report_footer(elements, styles, **meta)
    doc.build(elements)
    pdf_bytes = buffer.getvalue()
    buffer.close()
    return pdf_bytes


def generate_table_report_pdf(
    project,
    report_title,
    headers,
    rows,
    *,
    exported_by=None,
    verification_token=None,
    verify_url=None,
    section_title=None,
):
    """Branded PDF for arbitrary tabular data."""
    meta = _meta_kwargs(exported_by, verification_token, verify_url)
    buffer = io.BytesIO()
    section_title = section_title or report_title

    if project:
        doc, header, styles = _create_base_doc(buffer, project, report_title, exported_by=exported_by)
        elements = [header, Spacer(1, 24)]
    else:
        doc = SimpleDocTemplate(
            buffer, pagesize=letter, rightMargin=40, leftMargin=40, topMargin=40, bottomMargin=40,
        )
        styles = getSampleStyleSheet()
        header_style = ParagraphStyle(
            'GenericTitle', parent=styles['Heading1'], fontName='Helvetica-Bold',
            fontSize=20, textColor=colors.HexColor("#0f172a"), spaceAfter=12,
        )
        elements = [Paragraph(f"<b>BuildWise</b> — {report_title}", header_style), Spacer(1, 16)]

    section_style = ParagraphStyle(
        'TableSection', parent=styles['Heading2'], fontName='Helvetica-Bold',
        fontSize=16, textColor=colors.HexColor("#1e293b"), spaceAfter=10, spaceBefore=4,
    )
    elements.append(Paragraph(section_title, section_style))

    table_data = _wrap_table_rows(headers, rows, styles)

    if len(table_data) > 1:
        col_count = len(headers)
        page_width = letter[0] - 80
        col_width = max(48, page_width / max(col_count, 1))
        t_table = Table(table_data, colWidths=[col_width] * col_count, repeatRows=1)
        t_table.setStyle(_get_table_style())
        elements.append(t_table)
    else:
        elements.append(Paragraph("No data rows in this report.", styles['Normal']))

    append_report_footer(elements, styles, **meta)
    doc.build(elements)
    pdf_bytes = buffer.getvalue()
    buffer.close()
    return pdf_bytes


def generate_table_report_excel(report_title, headers, rows):
    from openpyxl import Workbook
    from openpyxl.styles import Alignment

    wb = Workbook()
    ws = wb.active
    ws.title = (report_title or "Report")[:31]
    ws.append(list(headers))
    for row in rows:
        ws.append([str(cell) if cell is not None else "" for cell in row])

    wrap = Alignment(wrap_text=True, vertical="top")
    for row in ws.iter_rows(min_row=1, max_row=ws.max_row, min_col=1, max_col=len(headers)):
        for cell in row:
            cell.alignment = wrap

    buffer = io.BytesIO()
    wb.save(buffer)
    buffer.seek(0)
    return buffer.getvalue()

